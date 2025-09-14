"""
Flask приложение для анализа классов SiTex
"""
import os
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, make_response
from data_service import DataService
from config import config
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)

# Увеличиваем лимиты для больших JSON ответов
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB
app.config['JSON_AS_ASCII'] = False

# Добавляем фильтр для форматирования даты
@app.template_filter('strftime')
def strftime_filter(date_str, format='%d.%m.%Y %H:%M'):
    if date_str:
        # Убираем время зоны и миллисекунды если есть
        date_str = date_str.split('+')[0].split('.')[0]
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
            return date_obj.strftime(format)
        except:
            return date_str.replace('T', ' ')[:16]
    return '-'

print(os.environ.get('JAVA_HOME'))
# Инициализация сервиса данных
data_service = DataService()

@app.route('/')
def index():
    """Главная страница со списком классов"""
    
    # Получаем параметры из запроса
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    # Получаем данные
    result = data_service.get_classes(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None,
        source_base_url=source_base_url if source_base_url else None
    )
    
    # Получаем статистику
    statistics = data_service.get_statistics()
    
    return render_template('index.html', 
                         result=result, 
                         statistics=statistics,
                         current_filters={
                             'search': search,
                             'status_variance': status_variance,
                             'event': event,
                             'a_priznak': a_priznak,
                             'base_url': base_url,
                             'source_base_url': source_base_url
                         })

@app.route('/groups')
def groups():
    """Страница со списком групп атрибутов"""
    
    # Получаем параметры из запроса
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    # Получаем данные
    result = data_service.get_groups(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None,
        source_base_url=source_base_url if source_base_url else None
    )
    
    # Получаем статистику
    statistics = data_service.get_statistics()
    
    return render_template('groups.html', 
                         result=result, 
                         statistics=statistics,
                         current_filters={
                             'search': search,
                             'status_variance': status_variance,
                             'event': event,
                             'a_priznak': a_priznak,
                             'base_url': base_url,
                             'source_base_url': source_base_url
                         })

@app.route('/attributes')
def attributes():
    """Страница со списком атрибутов"""
    
    # Получаем параметры из запроса
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений применяются только в режиме анализа исключений
    if analyze_exceptions:
        source_target_filter = request.args.get('source_target_filter', '')
        property_filter = request.args.getlist('property_filter')  # Множественный выбор свойств
        
        # Обработка чекбокса: если есть "true" в списке значений, то True, иначе False
        show_update_actions_values = request.args.getlist('show_update_actions')
        show_update_actions = 'true' in show_update_actions_values
    else:
        # В быстром режиме игнорируем фильтры исключений
        source_target_filter = None
        property_filter = None  
        show_update_actions = True
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    # Получаем данные
    result = data_service.get_attributes(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None,
        source_base_url=source_base_url if source_base_url else None,
        exception_action_filter=exception_action_filter,
        analyze_exceptions=analyze_exceptions,
        source_target_filter=source_target_filter,
        property_filter=property_filter,
        show_update_actions=show_update_actions
    )
    
    # Получаем статистику
    statistics = data_service.get_statistics()
    
    return render_template('attributes.html', 
                         result=result, 
                         statistics=statistics,
                         current_filters={
                             'search': search,
                             'status_variance': status_variance,
                             'event': event,
                             'a_priznak': a_priznak,
                             'base_url': base_url,
                             'source_base_url': source_base_url,
                             'exception_action_filter': exception_action_filter,
                             'analyze_exceptions': analyze_exceptions,
                             'source_target_filter': source_target_filter,
                             'property_filter': property_filter,
                             'show_update_actions': show_update_actions
                         })

@app.route('/classes')
def classes():
    """Страница с анализом классов по исключениям"""
    
    # Получаем параметры из запроса
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений применяются только в режиме анализа исключений
    if analyze_exceptions:
        source_target_filter = request.args.get('source_target_filter', '')
        property_filter = request.args.getlist('property_filter')  # Множественный выбор свойств
        
        # Обработка чекбокса: если есть "true" в списке значений, то True, иначе False
        show_update_actions_values = request.args.getlist('show_update_actions')
        show_update_actions = 'true' in show_update_actions_values
    else:
        # В быстром режиме игнорируем фильтры исключений
        source_target_filter = None
        property_filter = None  
        show_update_actions = True
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    # Получаем данные с анализом исключений
    result = data_service.get_classes_with_exceptions(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None,
        source_base_url=source_base_url if source_base_url else None,
        exception_action_filter=exception_action_filter,
        analyze_exceptions=analyze_exceptions,
        source_target_filter=source_target_filter,
        property_filter=property_filter,
        show_update_actions=show_update_actions
    )
    
    # Получаем статистику
    statistics = data_service.get_statistics()
    
    return render_template('classes.html', 
                         result=result, 
                         statistics=statistics,
                         current_filters={
                             'search': search,
                             'status_variance': status_variance,
                             'event': event,
                             'a_priznak': a_priznak,
                             'base_url': base_url,
                             'source_base_url': source_base_url,
                             'exception_action_filter': exception_action_filter,
                             'analyze_exceptions': analyze_exceptions,
                             'source_target_filter': source_target_filter,
                             'property_filter': property_filter,
                             'show_update_actions': show_update_actions
                         })

@app.route('/class/<int:class_ouid>')
def class_detail(class_ouid):
    """Детальная страница класса"""
    
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    
    result = data_service.get_class_details(
        class_ouid, 
        base_url if base_url else None,
        source_base_url if source_base_url else None,
        search if search else None,
        status_variance,
        event,
        a_priznak
    )
    
    if 'error' in result:
        return render_template('error.html', error=result['error'])
    
    # Добавляем текущие фильтры для отображения на странице
    result['current_filters'] = {
        'search': search,
        'status_variance': status_variance,
        'event': event,
        'a_priznak': a_priznak,
        'base_url': base_url,
        'source_base_url': source_base_url
    }
    
    return render_template('class_detail.html', data=result)

@app.route('/api/classes')
def api_classes():
    """API для получения списка классов (для AJAX)"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    result = data_service.get_classes(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None
    )
    
    return jsonify(result)

@app.route('/api/groups')
def api_groups():
    """API для получения списка групп (для AJAX)"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    result = data_service.get_groups(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None
    )
    
    return jsonify(result)

@app.route('/api/attributes')
def api_attributes():
    """API для получения списка атрибутов (для AJAX)"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    result = data_service.get_attributes(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None
    )
    
    return jsonify(result)

@app.route('/api/classes_with_exceptions')
def api_classes_with_exceptions():
    """API для получения списка классов с анализом исключений (для AJAX)"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений
    source_target_filter = request.args.get('source_target_filter', '')
    property_filter = request.args.getlist('property_filter')
    show_update_actions_values = request.args.getlist('show_update_actions')
    show_update_actions = 'true' in show_update_actions_values
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 20
    
    result = data_service.get_classes_with_exceptions(
        page=page,
        per_page=per_page,
        search=search if search else None,
        status_variance=status_variance,
        event=event,
        a_priznak=a_priznak,
        base_url=base_url if base_url else None,
        source_base_url=source_base_url if source_base_url else None,
        exception_action_filter=exception_action_filter,
        analyze_exceptions=analyze_exceptions,
        source_target_filter=source_target_filter if source_target_filter else None,
        property_filter=property_filter if property_filter else None,
        show_update_actions=show_update_actions
    )
    
    return jsonify(result)

@app.route('/api/class/<int:class_ouid>')
def api_class_detail(class_ouid):
    """API для получения деталей класса (для AJAX)"""
    
    base_url = request.args.get('base_url', '')
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    
    result = data_service.get_class_details(
        class_ouid, 
        base_url if base_url else None,
        search if search else None,
        status_variance,
        event,
        a_priznak
    )
    return jsonify(result)

@app.route('/api/statistics')
def api_statistics():
    """API для получения статистики"""
    
    result = data_service.get_statistics()
    return jsonify(result)

# ===== API для работы с исключениями =====

@app.route('/exceptions')
def exceptions():
    """Страница управления исключениями"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    entity_type = request.args.get('entity_type', '')
    search = request.args.get('search', '')
    
    # Проверяем корректность значений
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 200:
        per_page = 50
    
    # Получаем данные исключений
    result = data_service.get_exceptions(
        page=page,
        per_page=per_page,
        entity_type=entity_type if entity_type else None,
        search=search if search else None
    )
    
    return render_template('exceptions.html', 
                         result=result,
                         current_filters={
                             'entity_type': entity_type,
                             'search': search
                         })

@app.route('/api/exceptions')
def api_exceptions():
    """API для получения списка исключений (для AJAX)"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    entity_type = request.args.get('entity_type', '')
    search = request.args.get('search', '')
    
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 200:
        per_page = 50
    
    result = data_service.get_exceptions(
        page=page,
        per_page=per_page,
        entity_type=entity_type if entity_type else None,
        search=search if search else None
    )
    
    return jsonify(result)

@app.route('/api/exceptions/<int:exception_id>')
def api_get_exception(exception_id):
    """API для получения исключения по ID"""
    
    result = data_service.get_exception(exception_id)
    return jsonify(result)

@app.route('/api/exceptions', methods=['POST'])
def api_create_exception():
    """API для создания нового исключения"""
    
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Отсутствуют данные"}), 400
    
    entity_type = data.get('entity_type')
    entity_name = data.get('entity_name')
    property_name = data.get('property_name')
    action = data.get('action', 0)
    
    if not entity_type or not entity_name or not property_name:
        return jsonify({"error": "Обязательные поля: entity_type, entity_name, property_name"}), 400
    
    if entity_type not in ['class', 'group', 'attribute']:
        return jsonify({"error": "entity_type должен быть: class, group или attribute"}), 400
    
    if action not in [0, 2]:
        return jsonify({"error": "action должен быть: 0 (игнорировать) или 2 (обновить)"}), 400
    
    result = data_service.create_exception(entity_type, entity_name, property_name, action)
    
    if "error" in result:
        return jsonify(result), 400
    else:
        return jsonify(result), 201

@app.route('/api/exceptions/<int:exception_id>', methods=['PUT'])
def api_update_exception(exception_id):
    """API для обновления исключения"""
    
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Отсутствуют данные"}), 400
    
    entity_type = data.get('entity_type')
    entity_name = data.get('entity_name') 
    property_name = data.get('property_name')
    action = data.get('action')
    
    # Валидация если поля переданы
    if entity_type and entity_type not in ['class', 'group', 'attribute']:
        return jsonify({"error": "entity_type должен быть: class, group или attribute"}), 400
    
    if action is not None and action not in [0, 2]:
        return jsonify({"error": "action должен быть: 0 (игнорировать) или 2 (обновить)"}), 400
    
    result = data_service.update_exception(
        exception_id, entity_type, entity_name, property_name, action
    )
    
    if "error" in result:
        return jsonify(result), 400
    else:
        return jsonify(result)

@app.route('/api/exceptions/<int:exception_id>', methods=['DELETE'])
def api_delete_exception(exception_id):
    """API для удаления исключения"""
    
    result = data_service.delete_exception(exception_id)
    
    if "error" in result:
        return jsonify(result), 400
    else:
        return jsonify(result)

@app.route('/api/class/<int:class_ouid>/load-actions', methods=['POST'])
def api_load_actions(class_ouid):
    """API для загрузки действий из списка исключений"""
    
    # Получаем параметры фильтрации из тела запроса
    filters = request.get_json() or {}
    search = filters.get('search', '')
    status_variance = filters.get('status_variance', '')
    event = filters.get('event', '')
    
    # Конвертируем пустые строки в None, числовые значения в int
    search = search if search else None
    status_variance = int(status_variance) if status_variance else None
    event = int(event) if event else None
    
    result = data_service.load_actions_from_exceptions(class_ouid, search, status_variance, event)
    
    if "error" in result:
        return jsonify(result), 400
    else:
        return jsonify(result)

@app.route('/api/class/<int:class_ouid>/save-actions', methods=['POST'])
def api_save_actions(class_ouid):
    """API для записи действий в БД"""
    
    # Получаем параметры фильтрации из тела запроса
    filters = request.get_json() or {}
    search = filters.get('search', '')
    status_variance = filters.get('status_variance', '')
    event = filters.get('event', '')
    
    # Конвертируем пустые строки в None, числовые значения в int
    search = search if search else None
    status_variance = int(status_variance) if status_variance else None
    event = int(event) if event else None
    
    result = data_service.save_actions_to_db(class_ouid, search, status_variance, event)
    
    if "error" in result:
        return jsonify(result), 400
    else:
        return jsonify(result)

@app.route('/api/migrate-actions', methods=['POST'])
def api_migrate_actions():
    """API для миграции действий с -1 на 2"""
    result = data_service.migrate_actions_from_minus_one_to_two()
    
    if "error" in result:
        return jsonify(result), 400
    else:
        return jsonify(result), 200



@app.route('/api/reload-exceptions', methods=['POST'])
def api_reload_exceptions():
    """API для принудительной перезагрузки данных исключений из файлов"""
    
    try:
        # Принудительно перезагружаем данные исключений
        result = data_service.db_manager.init_exceptions_data(force_reload=True)
        
        if result:
            return jsonify({
                "success": True,
                "message": "Данные исключений успешно перезагружены из файлов"
            })
        else:
            return jsonify({"error": "Ошибка перезагрузки данных исключений"}), 500
            
    except Exception as e:
        return jsonify({"error": f"Ошибка перезагрузки исключений: {e}"}), 500

# ===== Эндпоинты для экспорта в Excel =====

@app.route('/export/classes.xlsx')
def export_classes_xlsx():
    """Экспорт классов в Excel с учетом фильтров"""
    
    # Получаем все те же параметры что и для обычного просмотра
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений применяются только в режиме анализа исключений
    if analyze_exceptions:
        source_target_filter = request.args.get('source_target_filter', '')
        property_filter = request.args.getlist('property_filter')
        show_update_actions_values = request.args.getlist('show_update_actions')
        show_update_actions = 'true' in show_update_actions_values
    else:
        source_target_filter = None
        property_filter = None  
        show_update_actions = True
    
    try:
        # Получаем данные без пагинации (устанавливаем per_page = 100000)
        result = data_service.get_classes_with_exceptions(
            page=1,
            per_page=100000,  # Получаем все записи
            search=search if search else None,
            status_variance=status_variance,
            event=event,
            a_priznak=a_priznak,
            base_url=base_url if base_url else None,
            source_base_url=source_base_url if source_base_url else None,
            exception_action_filter=exception_action_filter,
            analyze_exceptions=analyze_exceptions,
            source_target_filter=source_target_filter,
            property_filter=property_filter,
            show_update_actions=show_update_actions
        )
        
        if 'error' in result:
            return jsonify({"error": result['error']}), 500
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Классы"
        
        # Определяем заголовки в зависимости от режима
        if analyze_exceptions and result.get('classes_by_action'):
            # Полный режим с анализом исключений
            headers = ['ID', 'Имя', 'Описание', 'Свойство', 'Source', 'Target', 'Признак', 'Действие']
            ws.append(headers)
            
            # Стилизуем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Добавляем данные для обновления
            if result['classes_by_action'].get('update_list'):
                for class_item in result['classes_by_action']['update_list']:
                    priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(class_item.get('a_priznak'), 'Не определен')
                    ws.append([
                        class_item.get('ouid', ''),
                        class_item.get('name', ''),
                        class_item.get('description', ''),
                        class_item.get('property_name', ''),
                        class_item.get('source', ''),
                        class_item.get('target', ''),
                        priznak_text,
                        'Обновить'
                    ])
            
            # Добавляем данные для игнорирования
            if result['classes_by_action'].get('ignore_list'):
                for class_item in result['classes_by_action']['ignore_list']:
                    priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(class_item.get('a_priznak'), 'Не определен')
                    ws.append([
                        class_item.get('ouid', ''),
                        class_item.get('name', ''),
                        class_item.get('description', ''),
                        class_item.get('property_name', ''),
                        class_item.get('source', ''),
                        class_item.get('target', ''),
                        priznak_text,
                        'Игнорировать'
                    ])
            
            # Добавляем данные без действий
            if result['classes_by_action'].get('no_action_list'):
                for class_item in result['classes_by_action']['no_action_list']:
                    priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(class_item.get('a_priznak'), 'Не определен')
                    ws.append([
                        class_item.get('ouid', ''),
                        class_item.get('name', ''),
                        class_item.get('description', ''),
                        class_item.get('property_name', ''),
                        class_item.get('source', ''),
                        class_item.get('target', ''),
                        priznak_text,
                        'Без действия'
                    ])
        
        elif result.get('classes', {}).get('fast_mode'):
            # Быстрый режим
            headers = ['ID', 'Имя', 'Описание', 'Статус', 'Действие', 'Признак']
            ws.append(headers)
            
            # Стилизуем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Добавляем данные
            for class_item in result['classes']['fast_mode']:
                status_text = {0: 'Идентичны', 1: 'Отсутствует', 2: 'Отличаются'}.get(class_item.get('a_status_variance'), str(class_item.get('a_status_variance', '')))
                action_text = {0: 'Игнорировать', 1: 'Добавить', 2: 'Обновить'}.get(class_item.get('a_event'), str(class_item.get('a_event', '')))
                priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(class_item.get('a_priznak'), 'Не определен')
                
                ws.append([
                    class_item.get('ouid', ''),
                    class_item.get('name', ''),
                    class_item.get('description', ''),
                    status_text,
                    action_text,
                    priznak_text
                ])
        else:
            # Нет данных
            ws.append(['Нет данных для экспорта'])
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Генерируем имя файла с датой
        filename = f"classes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return jsonify({"error": f"Ошибка экспорта: {str(e)}"}), 500

@app.route('/export/attributes.xlsx')
def export_attributes_xlsx():
    """Экспорт атрибутов в Excel с учетом фильтров"""
    
    # Получаем все те же параметры что и для обычного просмотра
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений применяются только в режиме анализа исключений
    if analyze_exceptions:
        source_target_filter = request.args.get('source_target_filter', '')
        property_filter = request.args.getlist('property_filter')
        show_update_actions_values = request.args.getlist('show_update_actions')
        show_update_actions = 'true' in show_update_actions_values
    else:
        source_target_filter = None
        property_filter = None  
        show_update_actions = True
    
    try:
        # Получаем данные без пагинации (устанавливаем per_page = 100000)
        result = data_service.get_attributes(
            page=1,
            per_page=100000,  # Получаем все записи
            search=search if search else None,
            status_variance=status_variance,
            event=event,
            a_priznak=a_priznak,
            base_url=base_url if base_url else None,
            source_base_url=source_base_url if source_base_url else None,
            exception_action_filter=exception_action_filter,
            analyze_exceptions=analyze_exceptions,
            source_target_filter=source_target_filter,
            property_filter=property_filter,
            show_update_actions=show_update_actions
        )
        
        if 'error' in result:
            return jsonify({"error": result['error']}), 500
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Атрибуты"
        
        # Определяем заголовки в зависимости от режима
        if analyze_exceptions and result.get('classes'):
            # Полный режим с анализом исключений - группировка по классам
            headers = ['Класс', 'ID', 'Имя', 'Заголовок', 'Тип', 'Свойство', 'Source', 'Target', 'Признак', 'Действие']
            ws.append(headers)
            
            # Стилизуем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Добавляем данные по классам
            for class_name, class_data in result['classes'].items():
                # Атрибуты для обновления
                if class_data.get('attributes', {}).get('update_list'):
                    for attr in class_data['attributes']['update_list']:
                        priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(attr.get('a_priznak'), 'Не определен')
                        ws.append([
                            class_data.get('class_name', ''),
                            attr.get('ouid', ''),
                            attr.get('name', ''),
                            attr.get('title', ''),
                            attr.get('datatype_name', ''),
                            attr.get('property_name', ''),
                            attr.get('source', ''),
                            attr.get('target', ''),
                            priznak_text,
                            'Обновить'
                        ])
                
                # Атрибуты для игнорирования
                if class_data.get('attributes', {}).get('ignore_list'):
                    for attr in class_data['attributes']['ignore_list']:
                        priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(attr.get('a_priznak'), 'Не определен')
                        ws.append([
                            class_data.get('class_name', ''),
                            attr.get('ouid', ''),
                            attr.get('name', ''),
                            attr.get('title', ''),
                            attr.get('datatype_name', ''),
                            attr.get('property_name', ''),
                            attr.get('source', ''),
                            attr.get('target', ''),
                            priznak_text,
                            'Игнорировать'
                        ])
                
                # Атрибуты без действий
                if class_data.get('attributes', {}).get('no_action_list'):
                    for attr in class_data['attributes']['no_action_list']:
                        priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(attr.get('a_priznak'), 'Не определен')
                        ws.append([
                            class_data.get('class_name', ''),
                            attr.get('ouid', ''),
                            attr.get('name', ''),
                            attr.get('title', ''),
                            attr.get('datatype_name', ''),
                            attr.get('property_name', ''),
                            attr.get('source', ''),
                            attr.get('target', ''),
                            priznak_text,
                            'Без действия'
                        ])
        
        elif result.get('attributes', {}).get('fast_mode'):
            # Быстрый режим
            headers = ['ID', 'Имя', 'Заголовок', 'Класс', 'Тип', 'Признак']
            ws.append(headers)
            
            # Стилизуем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Добавляем данные
            for attr in result['attributes']['fast_mode']:
                priznak_text = {1: 'Переносим миграцией', 2: 'Не переносим', 3: 'Переносим не миграцией'}.get(attr.get('a_priznak'), 'Не определен')
                
                ws.append([
                    attr.get('ouid', ''),
                    attr.get('name', ''),
                    attr.get('title', ''),
                    attr.get('class_name', ''),
                    attr.get('datatype_name', ''),
                    priznak_text
                ])
        else:
            # Нет данных
            ws.append(['Нет данных для экспорта'])
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Генерируем имя файла с датой
        filename = f"attributes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        return jsonify({"error": f"Ошибка экспорта: {str(e)}"}), 500

@app.route('/api/generate_sql_scripts')
def generate_sql_scripts():
    """API для генерации SQL скриптов для отфильтрованных атрибутов (все записи)"""
    
    # Получаем все те же параметры что и для обычного просмотра
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений применяются только в режиме анализа исключений
    if analyze_exceptions:
        source_target_filter = request.args.get('source_target_filter', '')
        property_filter = request.args.getlist('property_filter')
        show_update_actions_values = request.args.getlist('show_update_actions')
        show_update_actions = 'true' in show_update_actions_values
    else:
        source_target_filter = None
        property_filter = None  
        show_update_actions = True
    
    try:
        # Получаем данные без пагинации (все записи)
        result = data_service.get_attributes(
            page=1,
            per_page=100000,  # Получаем все записи
            search=search if search else None,
            status_variance=status_variance,
            event=event,
            a_priznak=a_priznak,
            base_url=base_url if base_url else None,
            source_base_url=source_base_url if source_base_url else None,
            exception_action_filter=exception_action_filter,
            analyze_exceptions=analyze_exceptions,
            source_target_filter=source_target_filter,
            property_filter=property_filter,
            show_update_actions=show_update_actions
        )
        
        if 'error' in result:
            return jsonify({"error": result['error']}), 500
        
        # Генерируем SQL скрипты
        sql_scripts = []
        
        # Определяем значение A_EVENT (согласно требованиям пользователя - всегда 2)
        a_event_value = 2
        
        if analyze_exceptions and result.get('classes'):
            # Полный режим с анализом исключений - группировка по классам
            for class_name, class_data in result['classes'].items():
                # Атрибуты для обновления
                if class_data.get('attributes', {}).get('update_list'):
                    for attr in class_data['attributes']['update_list']:
                        where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                        if where_conditions:
                            script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                            sql_scripts.append(script)
                
                # Атрибуты для игнорирования
                if class_data.get('attributes', {}).get('ignore_list'):
                    for attr in class_data['attributes']['ignore_list']:
                        where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                        if where_conditions:
                            script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                            sql_scripts.append(script)
                
                # Атрибуты без действий
                if class_data.get('attributes', {}).get('no_action_list'):
                    for attr in class_data['attributes']['no_action_list']:
                        where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                        if where_conditions:
                            script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                            sql_scripts.append(script)
        
        elif result.get('attributes', {}).get('fast_mode'):
            # Быстрый режим - простая таблица атрибутов
            for attr in result['attributes']['fast_mode']:
                where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                if where_conditions:
                    script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                    sql_scripts.append(script)
        
        # Формируем итоговый результат
        all_scripts = '\n'.join(sql_scripts)
        
        return jsonify({
            "success": True,
            "scripts": all_scripts,
            "count": len(sql_scripts),
            "message": f"Сгенерировано {len(sql_scripts)} SQL скриптов"
        })
        
    except Exception as e:
        return jsonify({"error": f"Ошибка генерации скриптов: {str(e)}"}), 500

@app.route('/api/generate_sql_scripts_page')
def generate_sql_scripts_page():
    """API для генерации SQL скриптов только для текущей страницы"""
    
    # Получаем параметры включая пагинацию
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений применяются только в режиме анализа исключений
    if analyze_exceptions:
        source_target_filter = request.args.get('source_target_filter', '')
        property_filter = request.args.getlist('property_filter')
        show_update_actions_values = request.args.getlist('show_update_actions')
        show_update_actions = 'true' in show_update_actions_values
    else:
        source_target_filter = None
        property_filter = None  
        show_update_actions = True
    
    try:
        # Получаем данные с учетом пагинации (только текущая страница)
        result = data_service.get_attributes(
            page=page,
            per_page=per_page,  # Используем реальную пагинацию
            search=search if search else None,
            status_variance=status_variance,
            event=event,
            a_priznak=a_priznak,
            base_url=base_url if base_url else None,
            source_base_url=source_base_url if source_base_url else None,
            exception_action_filter=exception_action_filter,
            analyze_exceptions=analyze_exceptions,
            source_target_filter=source_target_filter,
            property_filter=property_filter,
            show_update_actions=show_update_actions
        )
        
        if 'error' in result:
            return jsonify({"error": result['error']}), 500
        
        # Генерируем SQL скрипты (та же логика что и в generate_sql_scripts)
        sql_scripts = []
        
        # Определяем значение A_EVENT (согласно требованиям пользователя - всегда 2)
        a_event_value = 2
        
        if analyze_exceptions and result.get('classes'):
            # Полный режим с анализом исключений - группировка по классам
            for class_name, class_data in result['classes'].items():
                # Атрибуты для обновления
                if class_data.get('attributes', {}).get('update_list'):
                    for attr in class_data['attributes']['update_list']:
                        where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                        if where_conditions:
                            script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                            sql_scripts.append(script)
                
                # Атрибуты для игнорирования
                if class_data.get('attributes', {}).get('ignore_list'):
                    for attr in class_data['attributes']['ignore_list']:
                        where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                        if where_conditions:
                            script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                            sql_scripts.append(script)
                
                # Атрибуты без действий
                if class_data.get('attributes', {}).get('no_action_list'):
                    for attr in class_data['attributes']['no_action_list']:
                        where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                        if where_conditions:
                            script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                            sql_scripts.append(script)
        
        elif result.get('attributes', {}).get('fast_mode'):
            # Быстрый режим - простая таблица атрибутов
            for attr in result['attributes']['fast_mode']:
                where_conditions = _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter)
                if where_conditions:
                    script = f"UPDATE SXATTR_SOURCE SET A_EVENT={a_event_value} WHERE {where_conditions};"
                    sql_scripts.append(script)
        
        # Формируем итоговый результат
        all_scripts = '\n'.join(sql_scripts)
        
        return jsonify({
            "success": True,
            "scripts": all_scripts,
            "count": len(sql_scripts),
            "message": f"Сгенерировано {len(sql_scripts)} SQL скриптов для текущей страницы"
        })
        
    except Exception as e:
        return jsonify({"error": f"Ошибка генерации скриптов: {str(e)}"}), 500

@app.route('/api/generate_data_update_scripts')
def generate_data_update_scripts():
    """API для генерации SQL скриптов обновления данных в целевой системе"""
    
    # Получаем те же параметры что и для обычных скриптов
    search = request.args.get('search', '')
    status_variance = request.args.get('status_variance', type=int)
    event = request.args.get('event', type=int)
    a_priznak = request.args.get('a_priznak', type=int)
    base_url = request.args.get('base_url', '')
    source_base_url = request.args.get('source_base_url', '')
    exception_action_filter = request.args.get('exception_action_filter', type=int)
    analyze_exceptions_param = request.args.get('analyze_exceptions', 'false').lower()
    analyze_exceptions = analyze_exceptions_param == 'true'
    
    # Фильтры исключений применяются только в режиме анализа исключений
    if analyze_exceptions:
        source_target_filter = request.args.get('source_target_filter', '')
        property_filter = request.args.getlist('property_filter')
        show_update_actions_values = request.args.getlist('show_update_actions')
        show_update_actions = 'true' in show_update_actions_values
    else:
        source_target_filter = None
        property_filter = None  
        show_update_actions = True
    
    # Проверяем что работаем только с source_to_null
    if source_target_filter != 'source_to_null':
        return jsonify({"error": "Скрипты обновления данных работают только с фильтром 'source_to_null'"}), 400
    
    if not property_filter:
        return jsonify({"error": "Для генерации скриптов обновления необходимо указать фильтр по свойствам"}), 400
    
    try:
        # Получаем данные с анализом исключений (все записи)
        result = data_service.get_attributes(
            page=1,
            per_page=100000,  # Получаем все записи
            search=search if search else None,
            status_variance=status_variance,
            event=event,
            a_priznak=a_priznak,
            base_url=base_url if base_url else None,
            source_base_url=source_base_url if source_base_url else None,
            exception_action_filter=exception_action_filter,
            analyze_exceptions=True,  # Принудительно включаем анализ для получения данных исключений
            source_target_filter=source_target_filter,
            property_filter=property_filter,
            show_update_actions=show_update_actions
        )
        
        if 'error' in result:
            return jsonify({"error": result['error']}), 500
        
        # Генерируем SQL скрипты обновления данных
        sql_scripts = []
        processed_count = 0
        
        if result.get('classes'):
            for class_name, class_data in result['classes'].items():
                # Обрабатываем все списки атрибутов
                for list_name in ['update_list', 'ignore_list', 'no_action_list']:
                    attr_list = class_data.get('attributes', {}).get(list_name, [])
                    for attr in attr_list:
                        attr_name = attr.get('name', '')
                        attr_ouid = attr.get('ouid', '')
                        exception_actions = attr.get('exception_actions', [])
                        
                        if not attr_name or not attr_ouid or not exception_actions:
                            continue
                        
                        # Ищем исключения с source_to_null для указанных свойств
                        for action in exception_actions:
                            prop_name = action.get('property_name', '')
                            source_val = action.get('source_value', '')
                            target_val = action.get('target_value', '')
                            
                            if (prop_name in property_filter and 
                                source_val and source_val.strip() != '' and source_val != 'null' and
                                (not target_val or target_val.strip() == '' or target_val == 'null')):
                                
                                # Извлекаем ID из source значения (например "11008462@SXAttrGrpSource" → "11008462")
                                source_id = _extract_id_from_source_value(source_val)
                                
                                if source_id:
                                    # Получаем маппинг поля БД для свойства
                                    field_mapping = _get_field_mapping(prop_name)
                                    
                                    if field_mapping:
                                        # Генерируем UPDATE скрипт
                                        update_script = _generate_update_script(
                                            attr_ouid, prop_name, source_id, field_mapping
                                        )
                                        
                                        if update_script:
                                            sql_scripts.append(update_script)
                                            processed_count += 1
        
        # Формируем итоговый результат
        all_scripts = '\n\n'.join(sql_scripts)
        
        # Отладочная информация о размере результата
        print(f"[DEBUG] Сгенерировано {len(sql_scripts)} скриптов, общий размер: {len(all_scripts)} символов")
        
        return jsonify({
            "success": True,
            "scripts": all_scripts,
            "count": len(sql_scripts),
            "processed_count": processed_count,
            "message": f"Сгенерировано {len(sql_scripts)} скриптов обновления данных ({len(all_scripts)} символов)"
        })
        
    except Exception as e:
        return jsonify({"error": f"Ошибка генерации скриптов обновления: {str(e)}"}), 500

def _extract_id_from_source_value(source_value):
    """Извлекает ID из строки типа '11008462@SXAttrGrpSource'"""
    try:
        if '@' in source_value:
            return source_value.split('@')[0].strip()
        return source_value.strip()
    except Exception:
        return None

def _get_field_mapping(property_name):
    """Получает маппинг поля БД для свойства через запрос к метаданным"""
    try:
        # Используем подключение к БД через data_service
        if not data_service.db_manager.connect():
            return None
        
        # Экранируем кавычки отдельно от f-string
        safe_property_name = property_name.replace("'", "''")
        query = f"SELECT map FROM sxattr WHERE name = '{safe_property_name}' LIMIT 1"
        result = data_service.db_manager.execute_query(query)
        
        if result and len(result) > 0:
            return result[0][0] 
        
        return None
    except Exception as e:
        print(f"[ERROR] Ошибка получения маппинга для {property_name}: {e}")
        return None
    finally:
        data_service.db_manager.disconnect()

def _generate_update_script(attr_ouid, property_name, source_id, field_mapping):
    """Генерирует UPDATE скрипт для обновления данных"""
    try:
        # Определяем исходную и целевую таблицы на основе свойства
        source_table = f"SXATTR_{property_name.upper()}_SOURCE"
        target_table = f"SXATTR_{property_name.upper()}"
        
        # Генерируем UPDATE скрипт
        script = f"""UPDATE sxattr
SET {field_mapping} = (
    SELECT grp.ouid
    FROM {target_table} grp
    JOIN SXOBJ o ON grp.ouid = o.ouid
    WHERE o.guid = (
        SELECT guid 
        FROM {source_table} sgrp 
        WHERE sgrp.ouid = {source_id} 
        LIMIT 1
    )
)
WHERE ouid = (
    SELECT A_LINK_TARGET 
    FROM SXATTR_SOURCE attrs 
    WHERE attrs.ouid = {attr_ouid} 
    LIMIT 1
); /* Обновление {property_name} для атрибута {attr_ouid} */"""
        
        return script
    except Exception as e:
        print(f"[ERROR] Ошибка генерации скрипта для {attr_ouid}: {e}")
        return None

def _build_where_conditions_for_update(attr, search, a_priznak, event, status_variance, property_filter, source_target_filter):
    """Построение WHERE условий для UPDATE скриптов с учетом всех фильтров"""
    conditions = []
    
    # DEBUG: Отладочная информация
    attr_name = attr.get('name', '')
    exception_actions = attr.get('exception_actions', [])
    print(f"[DEBUG] Building WHERE for attr '{attr_name}': property_filter={property_filter}, source_target_filter={source_target_filter}, exception_actions={len(exception_actions)} actions")
    if exception_actions:
        print(f"[DEBUG] Exception actions sample: {exception_actions[0]}")
    
    # Обязательное условие по имени атрибута
    if attr_name:
        safe_name = attr_name.replace("'", "''")
        conditions.append(f"NAME='{safe_name}'")
    
    # Добавляем основные фильтры если они были применены
    if a_priznak is not None:
        conditions.append(f"A_PRIZNAK={a_priznak}")
    
    if event is not None:
        conditions.append(f"A_EVENT={event}")
        
    if status_variance is not None:
        conditions.append(f"A_STATUS_VARIANCE={status_variance}")
    
    # Если есть связь с классом, добавляем условие по классу
    ouidsxclass = attr.get('ouidsxclass')
    if ouidsxclass:
        conditions.append(f"OUIDSXCLASS={ouidsxclass}")
    
    # Дополнительные условия на основе фильтров исключений
    exception_actions = attr.get('exception_actions', [])
    
    if not exception_actions:
        print(f"[DEBUG] No exception_actions found for '{attr_name}', will use A_LOG based filtering")
    
    # Фильтр по свойствам атрибутов
    # СЛУЧАЙ 1: Есть exception_actions - используем их
    if property_filter and exception_actions:
        # Извлекаем уникальные property_name из отфильтрованных исключений
        filtered_property_names = set()
        for action in exception_actions:
            prop_name = action.get('property_name', '')
            if prop_name and prop_name in property_filter:
                filtered_property_names.add(prop_name)
        
        print(f"[DEBUG] Property filter check for '{attr_name}': looking for {property_filter}, found {filtered_property_names}")
        
        if filtered_property_names:
            print(f"[DEBUG] Adding property filter for '{attr_name}': {filtered_property_names}")
            # Добавляем EXISTS условие для проверки наличия нужных свойств в A_LOG
            property_names_list = list(filtered_property_names)
            if len(property_names_list) == 1:
                # Одно свойство
                prop_name = property_names_list[0].replace("'", "''")
                conditions.append(f"A_LOG LIKE '%{prop_name}:%' /* Фильтр: только свойство {property_names_list[0]} */")
            else:
                # Несколько свойств - используем OR условие
                prop_conditions = []
                for prop_name in property_names_list:
                    safe_prop = prop_name.replace("'", "''")
                    prop_conditions.append(f"A_LOG LIKE '%{safe_prop}:%'")
                conditions.append(f"({' OR '.join(prop_conditions)}) /* Фильтр: свойства {', '.join(property_names_list)} */")
    
    # СЛУЧАЙ 2: Нет exception_actions, но есть фильтры - используем прямой поиск по A_LOG
    elif property_filter and not exception_actions:
        print(f"[DEBUG] Using direct A_LOG filtering for property_filter: {property_filter}")
        # Добавляем условие поиска по свойствам напрямую в A_LOG
        prop_conditions = []
        for prop_name in property_filter:
            safe_prop = prop_name.replace("'", "''")
            prop_conditions.append(f"A_LOG LIKE '%{safe_prop}%'")
        
        if len(prop_conditions) == 1:
            conditions.append(f"{prop_conditions[0]} /* Прямой фильтр свойства: {property_filter[0]} */")
        else:
            conditions.append(f"({' OR '.join(prop_conditions)}) /* Прямой фильтр свойств: {', '.join(property_filter)} */")
    
    # Фильтр по направлению изменений Source/Target
    if source_target_filter and exception_actions:
        # Собираем специфичные условия для source/target значений
        source_target_conditions = []
        
        for action in exception_actions:
            source_val = action.get('source_value', '')
            target_val = action.get('target_value', '')
            prop_name = action.get('property_name', '')
            
            # Логика фильтрации (повторяем из data_service.py)
            source_empty = not source_val or source_val.strip() == ''
            target_empty = not target_val or target_val.strip() == ''
            
            include_action = False
            
            if source_target_filter == 'source_to_null':
                # Source → Null (удаление): есть source, нет target
                include_action = not source_empty and target_empty
            elif source_target_filter == 'null_to_target':
                # Null → Target (добавление): нет source, есть target
                include_action = source_empty and not target_empty
            elif source_target_filter == 'source_to_target':
                # Source → Target (изменение): есть и source и target
                include_action = not source_empty and not target_empty
            elif source_target_filter == 'has_source':
                # Есть Source (любые с исходным значением)
                include_action = not source_empty
            elif source_target_filter == 'has_target':
                # Есть Target (любые с целевым значением)
                include_action = not target_empty
            
            if include_action and prop_name:
                # Экранируем значения для безопасности SQL
                safe_source = source_val.replace("'", "''") if source_val else ''
                safe_target = target_val.replace("'", "''") if target_val else ''
                safe_prop = prop_name.replace("'", "''")
                
                # Создаем условие проверяющее конкретное свойство и его значения в A_LOG
                if not source_empty and not target_empty:
                    # Есть и source и target
                    log_condition = f"A_LOG LIKE '%{safe_prop}: {safe_source} -> {safe_target}%'"
                elif not source_empty and target_empty:
                    # Только source, target пустой
                    log_condition = f"A_LOG LIKE '%{safe_prop}: {safe_source} ->%'"
                elif source_empty and not target_empty:
                    # Только target, source пустой
                    log_condition = f"A_LOG LIKE '%{safe_prop}:  -> {safe_target}%'"
                else:
                    # Оба пустые - общее условие по свойству
                    log_condition = f"A_LOG LIKE '%{safe_prop}:%'"
                
                source_target_conditions.append(log_condition)
        
        if source_target_conditions:
            print(f"[DEBUG] Adding source_target filter for '{attr_name}': {len(source_target_conditions)} conditions")
            # Объединяем условия через OR (любое из найденных сочетаний)
            if len(source_target_conditions) == 1:
                conditions.append(f"{source_target_conditions[0]} /* Фильтр {source_target_filter} */")
            else:
                conditions.append(f"({' OR '.join(source_target_conditions)}) /* Фильтр {source_target_filter}: {len(source_target_conditions)} условий */")
    
    # СЛУЧАЙ 2 для source_target_filter: Нет exception_actions, но есть фильтры - используем прямой поиск по A_LOG
    elif source_target_filter and not exception_actions:
        print(f"[DEBUG] Using direct A_LOG filtering for source_target_filter: {source_target_filter}")
        
        # Создаем условие для source_target_filter с учетом property_filter
        if property_filter:
            # Комбинируем фильтр свойства с фильтром направления
            combined_conditions = []
            for prop_name in property_filter:
                safe_prop = prop_name.replace("'", "''")
                
                if source_target_filter == 'source_to_null':
                    # Source → Null: ищем блок 'grp' где source имеет значение, а target = null
                    # Формат: "grp\n\tsource = value\n\ttarget = null"
                    log_condition = f"A_LOG LIKE '%{safe_prop}%source = %target = null%' AND A_LOG NOT LIKE '%{safe_prop}%source = null%'"
                elif source_target_filter == 'null_to_target':
                    # Null → Target: ищем блок 'grp' где source = null, а target имеет значение
                    # Формат: "grp\n\tsource = null\n\ttarget = value"
                    log_condition = f"A_LOG LIKE '%{safe_prop}%source = null%target = %' AND A_LOG NOT LIKE '%{safe_prop}%target = null%'"
                elif source_target_filter == 'source_to_target':
                    # Source → Target: оба поля имеют значения (НЕ null)
                    # Формат: "grp\n\tsource = value1\n\ttarget = value2"
                    log_condition = f"A_LOG LIKE '%{safe_prop}%source = %target = %' AND A_LOG NOT LIKE '%{safe_prop}%source = null%' AND A_LOG NOT LIKE '%{safe_prop}%target = null%'"
                elif source_target_filter == 'has_source':
                    # Есть Source: ищем блок 'grp' где source не равен null
                    log_condition = f"A_LOG LIKE '%{safe_prop}%source = %' AND A_LOG NOT LIKE '%{safe_prop}%source = null%'"
                elif source_target_filter == 'has_target':
                    # Есть Target: ищем блок 'grp' где target не равен null
                    log_condition = f"A_LOG LIKE '%{safe_prop}%target = %' AND A_LOG NOT LIKE '%{safe_prop}%target = null%'"
                else:
                    # Общий случай - просто проверяем наличие свойства в блоке
                    log_condition = f"A_LOG LIKE '%{safe_prop}%'"
                
                combined_conditions.append(log_condition)
            
            if len(combined_conditions) == 1:
                conditions.append(f"{combined_conditions[0]} /* Прямой фильтр {source_target_filter} для {property_filter[0]} */")
            else:
                conditions.append(f"({' OR '.join(combined_conditions)}) /* Прямой фильтр {source_target_filter} для свойств {', '.join(property_filter)} */")
        else:
            # Только source_target_filter без привязки к конкретным свойствам - сложнее реализовать
            print(f"[DEBUG] Warning: source_target_filter without property_filter is not supported for direct A_LOG filtering")
    
    final_where = " AND ".join(conditions)
    print(f"[DEBUG] Final WHERE condition for '{attr_name}': {final_where}")
    return final_where

@app.errorhandler(404)
def not_found(error):
    return render_template('error.html', error="Страница не найдена"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', error="Внутренняя ошибка сервера"), 500

if __name__ == '__main__':
    # Создаем директории для статики и шаблонов если их нет
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    
    print("🚀 Запуск Flask приложения...")
    print(f"📊 Подключение к PostgreSQL: {config.postgres.host}:{config.postgres.port}")
    print(f"🔗 Базовый URL админки: {config.sitex_context_url}")
    
    # Конфигурация из переменных окружения
    port = int(os.getenv('PORT', 5001))
    debug = os.getenv('FLASK_DEBUG', 'false').lower() in ('true', '1', 'yes')
    
    app.run(debug=debug, host='0.0.0.0', port=port) 